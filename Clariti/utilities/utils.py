import os
import allure
import openpyxl
from openpyxl.styles import PatternFill

from Clariti.utilities.XLUtils import getRowCount


def verify_status(testcasenumber, status):
    with allure.step("VERIFY " + testcasenumber + " STATUS"):
        if status:
            allure.attach("<h3 style='color:green;'>Test case PASSED</h3>", name=testcasenumber + " ✅ PASS",
                          attachment_type=allure.attachment_type.HTML)
        else:
            allure.attach("<h3 style='color:red;'>Test case FAILED</h3>", name=testcasenumber + " ❌ FAIL",
                          attachment_type=allure.attachment_type.HTML)


def status_report_update_to_xls(sheet, status, testcase_number):
    # status = "PASS"
    project_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    path = os.path.join(project_path, "ReportDocument", "AutomationCases.xlsx")
    rows = getRowCount(path, sheet)

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]

    for r in range(2, rows + 1):
        # for c in range(1,column):
        user = ws.cell(row=r, column=1).value  # Assuming 'Testcase number' is in the first column (column A)
        if user == testcase_number:
            if status:
                cell = ws.cell(row=r, column=5, value="PASS")
                cell.font = openpyxl.styles.Font(color="000000")  # Black font color
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00",
                                        fill_type="solid")  # Green background
            elif status is None:
                # Update status text as "PASS" with green background
                cell = ws.cell(row=r, column=5, value="FAIL")
                cell.font = openpyxl.styles.Font(color="000000")  # Black font color
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                        fill_type="solid")  # Red background
            else:
                cell = ws.cell(row=r, column=5, value="Status not updated correctly")

    wb.save(path)
